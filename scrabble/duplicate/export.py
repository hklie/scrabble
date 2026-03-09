"""export.py — Result export for Duplicate Scrabble (CSV, Excel, HTML, PNG)."""

import os

import pandas as pd


def build_results_dataframe(players, master_scores, total_rounds):
    """Build a results DataFrame.

    Rows: players sorted by total score descending.
    Columns: Round 1, Round 2, ..., Total, % vs Master.
    """
    master_total = sum(master_scores)

    # Sort players by total score descending
    sorted_players = sorted(players, key=lambda p: p.total_score, reverse=True)

    rows = []
    for i, player in enumerate(sorted_players):
        row = {'Player': f'Player {i + 1}'}
        for r in range(total_rounds):
            score = player.round_scores[r] if r < len(player.round_scores) else 0
            row[f'Round {r + 1}'] = score
        row['Total'] = player.total_score
        pct = (player.total_score / master_total * 100) if master_total > 0 else 0
        row['% vs Master'] = round(pct, 1)
        rows.append(row)

    # Add master row
    master_row = {'Player': 'Master'}
    for r in range(total_rounds):
        master_row[f'Round {r + 1}'] = master_scores[r] if r < len(master_scores) else 0
    master_row['Total'] = master_total
    master_row['% vs Master'] = 100.0
    rows.insert(0, master_row)

    return pd.DataFrame(rows)


def export_csv(players, master_scores, total_rounds, filepath):
    """Export results as UTF-8-sig CSV."""
    df = build_results_dataframe(players, master_scores, total_rounds)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df.to_csv(filepath, index=False, encoding='utf-8-sig')
    return filepath


def export_excel(players, master_scores, total_rounds, filepath):
    """Export results as .xlsx with formatting."""
    df = build_results_dataframe(players, master_scores, total_rounds)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        ws = writer.sheets['Results']

        from openpyxl.styles import Font, Alignment

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Bold Total column
        total_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'Total':
                total_col = idx
                break

        if total_col:
            for row in ws.iter_rows(min_row=2, min_col=total_col,
                                     max_col=total_col, max_row=ws.max_row):
                for cell in row:
                    cell.font = Font(bold=True)

        # Auto-width columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max_len + 3

    return filepath


def export_html(players, master_scores, total_rounds, filepath):
    """Export results as styled HTML table."""
    df = build_results_dataframe(players, master_scores, total_rounds)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    style = """
<style>
    body { font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f5f5f5; }
    h1 { color: #2c3e50; }
    table { border-collapse: collapse; width: 100%; background: white;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
    th { background: #2c3e50; color: white; padding: 10px 14px; text-align: center; }
    td { padding: 8px 14px; text-align: center; border-bottom: 1px solid #eee; }
    tr:first-child td { background: #e8f4f8; font-weight: bold; }
    .total { font-weight: bold; background: #f0f7e6; }
</style>
"""

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Duplicate Scrabble Results</title>{style}</head>
<body>
<h1>Duplicate Scrabble Results</h1>
<p>{total_rounds} rounds played</p>
{df.to_html(index=False, classes='results', escape=False)}
</body>
</html>
"""
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    return filepath


def export_graphical(players, master_scores, total_rounds, filepath):
    """Export cumulative score progression as PNG line chart."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    fig, ax = plt.subplots(figsize=(12, 6))

    rounds = list(range(1, total_rounds + 1))

    # Master cumulative scores
    master_cumulative = []
    running = 0
    for s in master_scores:
        running += s
        master_cumulative.append(running)
    ax.plot(rounds, master_cumulative, 'k-', linewidth=2.5, label='Master', marker='o',
            markersize=4)

    # Player cumulative scores
    sorted_players = sorted(players, key=lambda p: p.total_score, reverse=True)
    colors = plt.cm.tab10.colors
    for i, player in enumerate(sorted_players):
        cumulative = []
        running = 0
        for r in range(total_rounds):
            score = player.round_scores[r] if r < len(player.round_scores) else 0
            running += score
            cumulative.append(running)
        color = colors[i % len(colors)]
        ax.plot(rounds, cumulative, '-', linewidth=1.5, color=color,
                label=f'Player {i + 1}', marker='s', markersize=3)

    ax.set_xlabel('Round')
    ax.set_ylabel('Cumulative Score')
    ax.set_title('Duplicate Scrabble — Score Progression')
    ax.legend(loc='upper left')
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(filepath, dpi=150)
    plt.close()
    return filepath


def export_results(players, master_scores, total_rounds, fmt, base_path):
    """Dispatch to the appropriate export function.

    Returns the output filepath.
    """
    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(base_path, exist_ok=True)

    ext_map = {
        'csv': 'csv',
        'excel': 'xlsx',
        'html': 'html',
        'graphical': 'png',
    }
    ext = ext_map.get(fmt, 'csv')
    filepath = os.path.join(base_path, f'duplicate_{timestamp}.{ext}')

    exporters = {
        'csv': export_csv,
        'excel': export_excel,
        'html': export_html,
        'graphical': export_graphical,
    }

    exporter = exporters.get(fmt, export_csv)
    return exporter(players, master_scores, total_rounds, filepath)
